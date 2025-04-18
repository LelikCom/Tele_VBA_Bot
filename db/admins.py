"""
admins.py

Модуль для административных действий с базой данных PostgreSQL.
Содержит функции:
- Просмотр таблиц и колонок
- Выполнение произвольных SQL-запросов
- Экспорт результата в Excel
"""

import logging
from typing import List, Tuple
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from os import getenv
from dotenv import load_dotenv

from db.connection import get_db_connection


load_dotenv()
ADMIN_CHAT_ID = int(getenv("ADMIN_CHAT_ID", "0"))


async def get_all_table_names(user_id: int) -> List[str]:
    """
    Возвращает список всех таблиц в схеме `public`,
    скрывая `user_contacts_vba` для не-админов.

    Args:
        user_id (int): ID Telegram-пользователя, запрашивающего список.

    Returns:
        List[str]: Список названий таблиц.
    """
    query = """
        SELECT table_name
        FROM information_schema.tables
        WHERE table_schema = 'public' AND table_type = 'BASE TABLE'
        ORDER BY table_name
    """
    try:
        async with get_db_connection() as conn:
            records = await conn.fetch(query)
            tables = [r['table_name'] for r in records]

            # Фильтрация по роли
            if user_id != ADMIN_CHAT_ID:
                tables = [name for name in tables if name != 'user_contacts_vba']

            return tables
    except Exception as e:
        logging.error(f"Не удалось получить список таблиц: {e}")
        return []


async def get_table_columns(table: str) -> List[Tuple[str, str]]:
    """
    Возвращает список колонок таблицы с их типами.

    Args:
        table (str): Название таблицы.

    Returns:
        List[Tuple[str, str]]: Кортежи (имя колонки, тип данных).
    """
    query = """
        SELECT column_name, data_type
        FROM information_schema.columns
        WHERE table_name = $1
        ORDER BY ordinal_position
    """
    try:
        async with get_db_connection() as conn:
            records = await conn.fetch(query, table)
            return [(r['column_name'], r['data_type']) for r in records]
    except Exception as e:
        logging.error(f"Не удалось получить колонки таблицы '{table}': {e}")
        return []


async def execute_custom_sql_query(query: str, user_id: int) -> pd.DataFrame:
    """
    Выполняет произвольный SQL-запрос и возвращает результат как DataFrame.
    Только SELECT доступен для не-админов. Остальным — предупреждение.

    Args:
        query (str): SQL-запрос.
        user_id (int): Telegram user_id, инициатор запроса.

    Returns:
        pd.DataFrame: Результат запроса или сообщение об ограничении.
    """
    try:
        async with get_db_connection() as conn:
            sql = query.strip().lower()

            if user_id != ADMIN_CHAT_ID and not sql.startswith('select'):
                logging.warning(f"⛔ Пользователь {user_id} пытался выполнить запрещённый запрос: {sql}")
                return pd.DataFrame({"Ошибка": ["Не наглей, тебе доступен только SELECT"]})

            if sql.startswith('select'):
                records = await conn.fetch(query)
                if records:
                    columns = list(records[0].keys())
                    data = [tuple(r) for r in records]
                    return pd.DataFrame(data, columns=columns)
                else:
                    return pd.DataFrame()
            else:
                # Админу можно всё
                await conn.execute(query)
                return pd.DataFrame()

    except Exception as e:
        logging.error(f"Ошибка выполнения SQL-запроса: {e}")
        return pd.DataFrame({"Ошибка": [str(e)]})


def df_to_excel_bytes(df: pd.DataFrame) -> BytesIO:
    """
    Преобразует DataFrame в Excel-файл и возвращает его в виде байтового потока.

    Args:
        df (pd.DataFrame): Табличные данные.

    Returns:
        BytesIO: Поток с Excel-файлом.
    """
    wb = Workbook()
    ws = wb.active

    # Заголовки
    for col_num, column in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num, value=column)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Данные
    for row_num, row in enumerate(df.itertuples(index=False), 2):
        for col_num, value in enumerate(row, 1):
            val = str(value)
            if len(val) > 40:
                val = val[:40]
            cell = ws.cell(row=row_num, column=col_num, value=val)
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    # Автоширина
    for col_num, column in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_num)
        max_len = max(
            [len(str(df.columns[col_num - 1]))]
            + [len(str(val)) for val in df.iloc[:, col_num - 1].astype(str).values]
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

